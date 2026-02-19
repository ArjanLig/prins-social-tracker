# migrate_excel.py
"""One-time migration: import existing Excel data into SQLite."""

from openpyxl import load_workbook
from database import init_db, insert_posts, log_upload, DEFAULT_DB

EXCEL_FILE = "Social cijfers 2026 PRINS.xlsx"

def migrate():
    init_db()
    wb = load_workbook(EXCEL_FILE, read_only=True)

    # Facebook posts (tab "Facebook cijfers")
    if "Facebook cijfers" in wb.sheetnames:
        ws = wb["Facebook cijfers"]
        fb_posts = []
        for r in range(4, ws.max_row + 1):
            date_val = ws.cell(row=r, column=2).value
            if not date_val:
                continue
            if hasattr(date_val, "isoformat"):
                date_str = date_val.isoformat()
            else:
                date_str = str(date_val)[:10]
            fb_posts.append({
                "date": date_str,
                "type": ws.cell(row=r, column=3).value or "Post",
                "text": ws.cell(row=r, column=4).value or "",
                "views": ws.cell(row=r, column=5).value or 0,
                "reach": ws.cell(row=r, column=6).value or 0,
                "likes": ws.cell(row=r, column=7).value or 0,
                "comments": ws.cell(row=r, column=8).value or 0,
                "shares": ws.cell(row=r, column=9).value or 0,
                "clicks": ws.cell(row=r, column=10).value or 0,
                "source": "excel_migration",
            })
        if fb_posts:
            count = insert_posts(DEFAULT_DB, fb_posts, platform="facebook", page="prins")
            log_upload(DEFAULT_DB, EXCEL_FILE, "facebook", "prins", count)
            print(f"Facebook: {count} posts geimporteerd")

    # Instagram posts (tab "Instagram cijfers")
    if "Instagram cijfers" in wb.sheetnames:
        ws = wb["Instagram cijfers"]
        ig_posts = []
        for r in range(4, ws.max_row + 1):
            date_val = ws.cell(row=r, column=2).value
            if not date_val:
                continue
            if hasattr(date_val, "isoformat"):
                date_str = date_val.isoformat()
            else:
                date_str = str(date_val)[:10]
            ig_posts.append({
                "date": date_str,
                "type": ws.cell(row=r, column=9).value or "Post",
                "text": ws.cell(row=r, column=8).value or "",
                "reach": ws.cell(row=r, column=10).value or 0,
                "views": ws.cell(row=r, column=11).value or 0,
                "likes": ws.cell(row=r, column=12).value or 0,
                "comments": ws.cell(row=r, column=13).value or 0,
                "shares": ws.cell(row=r, column=14).value or 0,
                "clicks": ws.cell(row=r, column=15).value or 0,
                "source": "excel_migration",
            })
        if ig_posts:
            count = insert_posts(DEFAULT_DB, ig_posts, platform="instagram", page="prins")
            log_upload(DEFAULT_DB, EXCEL_FILE, "instagram", "prins", count)
            print(f"Instagram: {count} posts geimporteerd")

    wb.close()
    print("Migratie voltooid!")

if __name__ == "__main__":
    migrate()
