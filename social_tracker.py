#!/usr/bin/env python3
"""
Prins Petfoods Social Media Tracker
Combineert Facebook API data met handmatige CSV exports
"""

import os
import sys
from datetime import datetime
from pathlib import Path
import requests
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd

# Load environment variables
load_dotenv()

# Configuration
ENABLE_INSTAGRAM = False  # Instagram nog niet werkend
FB_API_VERSION = "v21.0"
FB_BASE_URL = f"https://graph.facebook.com/{FB_API_VERSION}"

# Environment variables
PRINS_TOKEN = os.getenv("PRINS_TOKEN")
PRINS_PAGE_ID = os.getenv("PRINS_PAGE_ID")
PRINS_CSV_PATH = os.getenv("PRINS_CSV_PATH", "")  # Optioneel: pad naar Facebook CSV export

def validate_env():
    """Check if required environment variables are set"""
    missing = []
    if not PRINS_TOKEN:
        missing.append("PRINS_TOKEN")
    if not PRINS_PAGE_ID:
        missing.append("PRINS_PAGE_ID")
    
    if missing:
        print("Ontbrekende environment variabelen:")
        for var in missing:
            print(f"  - {var}")
        sys.exit(1)

def api_get(object_id, token, params=None):
    """Generic Facebook Graph API GET request"""
    url = f"{FB_BASE_URL}/{object_id}"
    params = params or {}
    params["access_token"] = token
    
    resp = requests.get(url, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()

def fetch_fb_page_stats(page_id, token):
    """Fetch Facebook page statistics (followers, fans)"""
    try:
        data = api_get(page_id, token, {"fields": "name,fan_count,followers_count"})
        return {
            "name": data.get("name"),
            "fans": data.get("fan_count", 0),
            "followers": data.get("followers_count", 0)
        }
    except Exception as e:
        print(f"  ⚠️  Fout bij ophalen page stats: {e}")
        return None

def load_facebook_csv(csv_path):
    """Load Facebook posts from CSV export"""
    if not csv_path or not Path(csv_path).exists():
        print(f"  ℹ️  Geen Facebook CSV gevonden")
        return None
    
    try:
        # Lees Facebook export (UTF-8 with BOM)
        df = pd.read_csv(csv_path, encoding='utf-8-sig')
        
        # Simplify column names
        df = df.rename(columns={
            'Titel': 'message',
            'Publicatietijdstip': 'datetime',
            'Datum': 'date',
            'Berichttype': 'type',
            'Weergaven': 'views',
            'Bereik': 'reach',
            'Reacties': 'likes',  # Facebook noemt het "Reacties" maar we gebruiken likes voor consistentie
            'Opmerkingen': 'comments',
            'Deelacties': 'shares',
            'Totaal aantal klikken': 'clicks'
        })
        
        # Selecteer alleen relevante kolommen
        columns_to_keep = ['message', 'date', 'datetime', 'type', 'views', 'reach', 
                          'likes', 'comments', 'shares', 'clicks']
        df = df[[col for col in columns_to_keep if col in df.columns]]
        
        print(f"  ✓ {len(df)} Facebook posts ingeladen uit CSV")
        return df
    
    except Exception as e:
        print(f"  ⚠️  Fout bij laden CSV: {e}")
        return None

def write_to_excel(prins_stats, prins_posts_df):
    """Write all data to Excel file"""
    excel_path = "Social cijfers 2026 PRINS.xlsx"
    
    # Check if file exists
    if Path(excel_path).exists():
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    
    # === Facebook KPI's Sheet ===
    if 'Facebook KPIs' not in wb.sheetnames:
        wb.create_sheet('Facebook KPIs')
    
    ws_fb_kpi = wb['Facebook KPIs']
    ws_fb_kpi.delete_rows(1, ws_fb_kpi.max_row)
    
    # Header
    ws_fb_kpi['A1'] = "Facebook KPI's"
    ws_fb_kpi['A1'].font = Font(bold=True, size=14)
    
    # Maanden
    months = ['Januari', 'Februari', 'Maart', 'April', 'Mei', 'Juni', 
              'Juli', 'Augustus', 'September', 'Oktober', 'November', 'December']
    for i, month in enumerate(months, start=2):
        ws_fb_kpi.cell(2, i).value = month
        ws_fb_kpi.cell(2, i).font = Font(bold=True)
    
    # Metrics rijen
    metrics = ['Fans', 'Volgers', 'Weergaven', 'Bereik', 'Engagement', 'Aantal posts']
    for i, metric in enumerate(metrics, start=3):
        ws_fb_kpi.cell(i, 1).value = metric
        ws_fb_kpi.cell(i, 1).font = Font(bold=True)
    
    # Vul huidige maand fans/volgers in
    current_month_col = datetime.now().month + 1  # +1 omdat kolom 1 is labels
    if prins_stats:
        ws_fb_kpi.cell(3, current_month_col).value = prins_stats['fans']  # Fans
        ws_fb_kpi.cell(4, current_month_col).value = prins_stats['followers']  # Volgers
    
    print("  ✓ Facebook KPIs geschreven")
    
    # === Facebook Cijfers Sheet ===
    if prins_posts_df is not None and not prins_posts_df.empty:
        if 'Facebook cijfers' not in wb.sheetnames:
            wb.create_sheet('Facebook cijfers')
        
        ws_fb_posts = wb['Facebook cijfers']
        ws_fb_posts.delete_rows(1, ws_fb_posts.max_row)
        
        # Header
        ws_fb_posts['A1'] = "Facebook resultaten per post"
        ws_fb_posts['A1'].font = Font(bold=True, size=14)
        ws_fb_posts['A2'] = "Posts"
        ws_fb_posts['A2'].font = Font(bold=True)
        
        # Column headers
        headers = ['#', 'Datum', 'Type post', 'Omschrijving', 'Weergaven', 'Bereik', 
                   'Likes', 'Reacties', 'Shares', 'Klikken', 'Totaal engagement']
        for col, header in enumerate(headers, start=1):
            cell = ws_fb_posts.cell(3, col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Data rows
        for idx, row in prins_posts_df.iterrows():
            row_num = idx + 4
            ws_fb_posts.cell(row_num, 1).value = idx + 1  # #
            ws_fb_posts.cell(row_num, 2).value = row.get('date', '')  # Datum
            ws_fb_posts.cell(row_num, 3).value = row.get('type', '')  # Type
            
            # Omschrijving (truncate lange teksten)
            message = row.get('message', '')
            if pd.notna(message):
                message_short = str(message)[:100] + '...' if len(str(message)) > 100 else str(message)
                ws_fb_posts.cell(row_num, 4).value = message_short
            
            ws_fb_posts.cell(row_num, 5).value = row.get('views', 0)  # Weergaven
            ws_fb_posts.cell(row_num, 6).value = row.get('reach', 0)  # Bereik
            ws_fb_posts.cell(row_num, 7).value = row.get('likes', 0)  # Likes
            ws_fb_posts.cell(row_num, 8).value = row.get('comments', 0)  # Reacties
            ws_fb_posts.cell(row_num, 9).value = row.get('shares', 0)  # Shares
            ws_fb_posts.cell(row_num, 10).value = row.get('clicks', 0)  # Klikken
            
            # Totaal engagement
            engagement = (row.get('likes', 0) + row.get('comments', 0) + 
                         row.get('shares', 0))
            ws_fb_posts.cell(row_num, 11).value = engagement
        
        # Auto-width columns
        for col in ws_fb_posts.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws_fb_posts.column_dimensions[col_letter].width = adjusted_width
        
        print(f"  ✓ {len(prins_posts_df)} Facebook posts geschreven")
    
    # Save workbook
    wb.save(excel_path)
    print(f"\n✅ Opgeslagen in {excel_path}")
    return excel_path

def main():
    print("=== Prins Petfoods Social Media Tracker ===\n")
    
    # Validate environment
    validate_env()
    
    # Fetch Facebook page stats
    print("Facebook: Prins Petfoods...")
    prins_stats = fetch_fb_page_stats(PRINS_PAGE_ID, PRINS_TOKEN)
    
    if prins_stats:
        print(f"  ✓ {prins_stats['name']}: {prins_stats['fans']} fans, {prins_stats['followers']} volgers")
    
    # Load Facebook CSV if available
    print("\nFacebook posts uit CSV...")
    prins_posts_df = load_facebook_csv(PRINS_CSV_PATH)
    
    # Instagram placeholder
    if ENABLE_INSTAGRAM:
        print("\nInstagram: Niet geïmplementeerd")
    else:
        print("\nInstagram: Overgeslagen (ENABLE_INSTAGRAM = False)")
        print("  ⚠️  Zet ENABLE_INSTAGRAM = True zodra verificatie werkt")
    
    # Write to Excel
    print("\nSchrijven naar Excel...")
    excel_path = write_to_excel(prins_stats, prins_posts_df)
    
    # Summary
    print("\n" + "="*50)
    print("✅ Klaar!")
    if prins_posts_df is None:
        print("\n💡 Tip: Exporteer Facebook posts CSV en zet PRINS_CSV_PATH in .env")
        print("   Dan worden posts automatisch ingeladen!")

if __name__ == "__main__":
    main()
