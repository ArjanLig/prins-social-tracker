"""
Social Media Tracker voor Prins Petfoods & Edupet.
Importeert Facebook + Instagram statistieken via CSV (Meta Business Suite)
of via de Playwright scraper.
Schrijft data naar het Excel tracking sheet en genereert een AI-analyse.
"""

import argparse
import os
import re
from datetime import datetime, timezone

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from fb_scraper import scrape_fb_page_posts, has_session as has_fb_session

load_dotenv()

EXCEL_FILE = "Social cijfers 2026 PRINS.xlsx"

DAGEN_NL = {
    "Monday": "Maandag", "Tuesday": "Dinsdag", "Wednesday": "Woensdag",
    "Thursday": "Donderdag", "Friday": "Vrijdag", "Saturday": "Zaterdag",
    "Sunday": "Zondag",
}

MAAND_NL = {
    1: "Januari", 2: "Februari", 3: "Maart", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Augustus",
    9: "September", 10: "Oktober", 11: "November", 12: "December",
}

PASTEL_FILLS = [
    PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),  # lichtblauw
    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # lichtgroen
    PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # licht oranje
    PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),  # licht paars
    PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # lichtgeel
    PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),  # pastel indigo
]
NO_FILL = PatternFill(fill_type=None)


def color_rows_by_month(ws, data_start_row: int, maand_col: int, max_col: int):
    """Geef rijen afwisselende pastelkleuren per maand."""
    current_maand = None
    color_idx = -1
    for r in range(data_start_row, ws.max_row + 1):
        maand = ws.cell(row=r, column=maand_col).value
        if not maand:
            continue
        if maand != current_maand:
            current_maand = maand
            color_idx = (color_idx + 1) % len(PASTEL_FILLS)
        fill = PASTEL_FILLS[color_idx]
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = fill

MAAND_COL = {
    1: "B", 2: "C", 3: "D", 4: "E", 5: "F", 6: "G",
    7: "H", 8: "I", 9: "J", 10: "K", 11: "L", 12: "M",
}


def dagdeel(hour: int) -> str:
    if hour < 6:
        return "Nacht"
    if hour < 12:
        return "Ochtend"
    if hour < 18:
        return "Middag"
    return "Avond"



# ─── Excel schrijven ────────────────────────────────────────────────

def write_ig_posts(wb, posts: list[dict]):
    """Schrijf Instagram posts naar tab 'Instagram cijfers'."""
    ws = wb["Instagram cijfers"]

    # Bepaal volgende rij (na header rij 3)
    next_row = ws.max_row + 1
    if next_row <= 3:
        next_row = 4

    # Bepaal volgnummer en bestaande post-ID's
    last_num = 0
    existing_ids = set()
    for r in range(4, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, (int, float)):
            last_num = max(last_num, int(v))
        existing_id = ws.cell(row=r, column=18).value  # R = interne post-id
        if isinstance(existing_id, str) and existing_id.strip():
            existing_ids.add(existing_id.strip())

    # Normaliseer CSV-formaat naar API-formaat
    for post in posts:
        if "timestamp" not in post and "date" in post:
            post["timestamp"] = post["date"]
        if "like_count" not in post and "likes" in post:
            post["like_count"] = post["likes"]
        if "comments_count" not in post and "comments" in post:
            post["comments_count"] = post["comments"]
        if "media_type" not in post and "type" in post:
            post["media_type"] = post["type"]
        if "caption" not in post and "text" in post:
            post["caption"] = post["text"]
        if "reach" not in post:
            post["reach"] = 0
        if "impressions" not in post and "views" in post:
            post["impressions"] = post["views"]

    appended = 0
    for post in posts:
        post_id = (post.get("id") or post.get("permalink") or "").strip()
        if post_id and post_id in existing_ids:
            continue
        ts = datetime.fromisoformat(post["timestamp"].replace("+0000", "+00:00"))
        bereik = post.get("reach", 0) or 0
        weergaven = post.get("impressions", 0) or 0
        likes = post.get("like_count", 0) or 0
        reacties = post.get("comments_count", 0) or 0
        shares = 0  # Niet beschikbaar via IG API
        totaal_engagement = likes + reacties
        er = (totaal_engagement / bereik * 100) if bereik > 0 else 0

        row = next_row + appended
        ws.cell(row=row, column=1, value=last_num + appended + 1)       # #
        ws.cell(row=row, column=2, value=ts.strftime("%Y-%m-%d"))  # Datum
        ws.cell(row=row, column=3, value=DAGEN_NL.get(ts.strftime("%A"), ""))  # Dag
        ws.cell(row=row, column=4, value=ts.strftime("%H:%M"))  # Tijd
        ws.cell(row=row, column=5, value=dagdeel(ts.hour))    # Dagdeel
        # F=Thema, G=Campagne (handmatig)
        caption = (post.get("caption") or "")[:100]
        ws.cell(row=row, column=8, value=caption)             # H=Omschrijving
        ws.cell(row=row, column=9, value=post.get("media_type", ""))  # I=Type post
        ws.cell(row=row, column=10, value=bereik)             # J=Bereik
        ws.cell(row=row, column=11, value=weergaven)          # K=Weergaven
        ws.cell(row=row, column=12, value=likes)              # L=Likes
        ws.cell(row=row, column=13, value=reacties)           # M=Reacties
        ws.cell(row=row, column=14, value=shares)             # N=Shares
        # O=Website klikken (niet via basic API)
        ws.cell(row=row, column=16, value=totaal_engagement)  # P=Totaal engagement
        ws.cell(row=row, column=17, value=round(er, 2))       # Q=ER
        ws.cell(row=row, column=18, value=post_id)            # R=Post ID (dedupe)
        maand_str = f"{MAAND_NL[ts.month]} {ts.year}"
        ws.cell(row=row, column=19, value=maand_str)         # S=Maand
        if post_id:
            existing_ids.add(post_id)
        appended += 1

    # Backfill Maand voor bestaande rijen
    for r in range(4, next_row):
        if ws.cell(row=r, column=19).value is None:
            date_val = ws.cell(row=r, column=2).value
            if isinstance(date_val, str) and len(date_val) >= 7:
                try:
                    dt = datetime.fromisoformat(date_val)
                    ws.cell(row=r, column=19, value=f"{MAAND_NL[dt.month]} {dt.year}")
                except ValueError:
                    pass
            elif hasattr(date_val, "month"):
                ws.cell(row=r, column=19, value=f"{MAAND_NL[date_val.month]} {date_val.year}")

    # Header
    ws.cell(row=3, column=19, value="Maand")

    # Verberg kolom R (post ID's)
    ws.column_dimensions["R"].hidden = True

    # Kleur rijen per maand
    color_rows_by_month(ws, data_start_row=4, maand_col=19, max_col=19)



def write_fb_posts(wb, posts: list[dict]):
    """Schrijf Facebook posts (van scraper) naar tab 'Facebook cijfers'."""
    ws = wb["Facebook cijfers"]

    # Bepaal volgende rij (na header rij 3)
    next_row = ws.max_row + 1
    if next_row <= 3:
        next_row = 4

    # Bepaal volgnummer en bestaande post-ID's (kolom L = 12 voor dedupe)
    last_num = 0
    existing_ids = set()
    for r in range(4, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, (int, float)):
            last_num = max(last_num, int(v))
        existing_id = ws.cell(row=r, column=12).value  # L = post-id voor dedupe
        if isinstance(existing_id, str) and existing_id.strip():
            existing_ids.add(existing_id.strip())

    appended = 0
    for post in posts:
        post_id = post.get("id", "").strip()
        if post_id and post_id in existing_ids:
            continue

        date_str = post.get("date", "")
        if date_str:
            ts = datetime.fromisoformat(date_str)
        else:
            continue  # Geen datum = skip

        likes = post.get("likes", 0)
        comments = post.get("comments", 0)
        shares = post.get("shares", 0)
        totaal_engagement = likes + comments + shares
        text = post.get("text") or "(geen tekst)"
        # Opschonen: regelovergangen → spatie, dubbele spaties weg, trim
        text = re.sub(r'\s*\n\s*', ' | ', text)  # newlines → pipe separator
        text = re.sub(r'\s{2,}', ' ', text).strip()
        text = text[:200]

        row = next_row + appended
        ws.cell(row=row, column=1, value=last_num + appended + 1)   # A: #
        ws.cell(row=row, column=2, value=ts.strftime("%Y-%m-%d"))   # B: Datum
        ws.cell(row=row, column=3, value=post.get("type") or "Post")  # C: Type post
        ws.cell(row=row, column=4, value=text)                      # D: Omschrijving
        views = post.get("views", 0)
        if views:
            ws.cell(row=row, column=5, value=views)                # E: Weergaven (video)
        bereik = post.get("reach", 0)
        if bereik:
            ws.cell(row=row, column=6, value=bereik)               # F: Bereik
        ws.cell(row=row, column=7, value=likes)                     # G: Likes
        ws.cell(row=row, column=8, value=comments)                  # H: Reacties
        ws.cell(row=row, column=9, value=shares)                    # I: Shares
        klikken = post.get("clicks", 0)
        if klikken:
            ws.cell(row=row, column=10, value=klikken)             # J: Klikken
        ws.cell(row=row, column=11, value=totaal_engagement)        # K: Totaal engagement
        ws.cell(row=row, column=12, value=post_id)                  # L: Post ID (dedupe, verborgen)
        maand_str = f"{MAAND_NL[ts.month]} {ts.year}"
        ws.cell(row=row, column=13, value=maand_str)               # M: Maand

        if post_id:
            existing_ids.add(post_id)
        appended += 1

    # Backfill Maand voor bestaande rijen
    for r in range(4, next_row):
        if ws.cell(row=r, column=13).value is None:
            date_val = ws.cell(row=r, column=2).value
            if isinstance(date_val, str) and len(date_val) >= 7:
                try:
                    dt = datetime.fromisoformat(date_val)
                    ws.cell(row=r, column=13, value=f"{MAAND_NL[dt.month]} {dt.year}")
                except ValueError:
                    pass
            elif hasattr(date_val, "month"):
                ws.cell(row=r, column=13, value=f"{MAAND_NL[date_val.month]} {date_val.year}")

    # Header
    ws.cell(row=3, column=13, value="Maand")

    # Verberg kolom L (post ID's)
    ws.column_dimensions["L"].hidden = True

    # Kleur rijen per maand
    color_rows_by_month(ws, data_start_row=4, maand_col=13, max_col=13)

    return appended


def write_fb_kpis(wb, posts: list[dict]):
    """Schrijf maandelijkse Facebook KPI's naar tab 'Facebook KPIs'."""
    ws = wb["Facebook KPIs"]
    now = datetime.now(timezone.utc)
    col = MAAND_COL.get(now.month, "B")

    # Filter posts van deze maand
    month_posts = []
    for p in posts:
        date_str = p.get("date", "")
        if not date_str:
            continue
        ts = datetime.fromisoformat(date_str).astimezone(timezone.utc)
        if ts.month == now.month and ts.year == now.year:
            month_posts.append(p)

    total_engagement = sum(
        p.get("likes", 0) + p.get("comments", 0) + p.get("shares", 0)
        for p in month_posts
    )
    num_posts = len(month_posts)

    ws[f"{col}7"] = total_engagement        # Engagement
    ws[f"{col}8"] = num_posts               # Aantal posts



# ─── AI Analyse ─────────────────────────────────────────────────────

def collect_summary(prins_fb_posts: list[dict], edupet_fb_posts: list[dict],
                    ig_posts: list[dict]) -> str:
    """Verzamel alle opgehaalde data in een gestructureerde tekst-samenvatting."""
    now = datetime.now(timezone.utc)
    lines = [f"=== Social Media Data Prins Petfoods — {now.strftime('%d-%m-%Y')} ===\n"]

    def _fb_post_summary(posts: list[dict], label: str):
        """Voeg Facebook post-samenvatting toe voor een pagina."""
        if not posts:
            lines.append("- Posts: geen data")
            return
        lines.append(f"- Aantal recente posts: {len(posts)}")
        total_likes = sum(p.get("likes", 0) for p in posts)
        total_comments = sum(p.get("comments", 0) for p in posts)
        total_shares = sum(p.get("shares", 0) for p in posts)
        lines.append(f"- Totaal: {total_likes} likes, {total_comments} reacties, {total_shares} shares")

        # Top posts
        sorted_posts = sorted(posts, key=lambda p: p.get("likes", 0) + p.get("comments", 0), reverse=True)
        lines.append(f"\n### Top posts {label} (hoogste engagement):")
        for i, post in enumerate(sorted_posts[:5], 1):
            msg = (post.get("text") or "(geen tekst)")[:80]
            msg = msg.encode("utf-8", errors="replace").decode("utf-8").replace("\n", " ")
            ts = post.get("date", "")[:10]
            lines.append(f"  {i}. ({ts}): {post.get('likes', 0)} likes, "
                         f"{post.get('comments', 0)} reacties, {post.get('shares', 0)} shares — \"{msg}\"")

        if len(sorted_posts) > 5:
            lines.append(f"\n### Minst presterende posts {label}:")
            for i, post in enumerate(sorted_posts[-3:], 1):
                msg = (post.get("text") or "(geen tekst)")[:80]
                msg = msg.encode("utf-8", errors="replace").decode("utf-8").replace("\n", " ")
                ts = post.get("date", "")[:10]
                lines.append(f"  {i}. ({ts}): {post.get('likes', 0)} likes, "
                             f"{post.get('comments', 0)} reacties, {post.get('shares', 0)} shares — \"{msg}\"")

    # Facebook Prins
    lines.append("## Facebook — Prins Petfoods")
    _fb_post_summary(prins_fb_posts, "Prins")

    # Facebook Edupet
    lines.append("\n## Facebook — Edupet")
    _fb_post_summary(edupet_fb_posts, "Edupet")

    # Instagram Prins
    if ig_posts:
        lines.append("\n## Instagram — Prins Petfoods")

        # Normaliseer veldnamen (CSV gebruikt andere keys)
        for p in ig_posts:
            if "timestamp" not in p and "date" in p:
                p["timestamp"] = p["date"]
            if "like_count" not in p and "likes" in p:
                p["like_count"] = p["likes"]
            if "comments_count" not in p and "comments" in p:
                p["comments_count"] = p["comments"]
            if "media_type" not in p and "type" in p:
                p["media_type"] = p["type"]
            if "caption" not in p and "text" in p:
                p["caption"] = p["text"]

        # Filter deze maand
        month_posts = []
        for p in ig_posts:
            ts = datetime.fromisoformat(p["timestamp"].replace("+0000", "+00:00"))
            if ts.month == now.month and ts.year == now.year:
                month_posts.append(p)

        lines.append(f"- Posts deze maand: {len(month_posts)}")
        total_reach = sum(p.get("reach", 0) or 0 for p in month_posts)
        total_impressions = sum(p.get("impressions", 0) or 0 for p in month_posts)
        total_likes = sum(p.get("like_count", 0) or 0 for p in month_posts)
        total_comments = sum(p.get("comments_count", 0) or 0 for p in month_posts)
        total_eng = total_likes + total_comments
        er = (total_eng / total_reach * 100) if total_reach > 0 else 0
        lines.append(f"- Totaal bereik deze maand: {total_reach}")
        lines.append(f"- Totaal weergaven deze maand: {total_impressions}")
        lines.append(f"- Totaal likes: {total_likes}, reacties: {total_comments}")
        lines.append(f"- Engagement rate: {er:.2f}%%")

        # Top posts op engagement
        sorted_posts = sorted(month_posts, key=lambda p: (p.get("like_count", 0) or 0) + (p.get("comments_count", 0) or 0), reverse=True)
        if sorted_posts:
            lines.append("\n### Top posts (hoogste engagement):")
            for i, p in enumerate(sorted_posts[:5], 1):
                caption = (p.get("caption") or "(geen caption)")[:80]
                likes = p.get("like_count", 0) or 0
                comments = p.get("comments_count", 0) or 0
                reach = p.get("reach", 0) or 0
                ts = p.get("timestamp", "")[:10]
                mtype = p.get("media_type", "?")
                lines.append(f"  {i}. ({ts}, {mtype}) {likes} likes, {comments} reacties, bereik {reach} — \"{caption}\"")

        # Flop posts
        if len(sorted_posts) > 3:
            lines.append("\n### Minst presterende posts:")
            for i, p in enumerate(sorted_posts[-3:], 1):
                caption = (p.get("caption") or "(geen caption)")[:80]
                likes = p.get("like_count", 0) or 0
                comments = p.get("comments_count", 0) or 0
                reach = p.get("reach", 0) or 0
                ts = p.get("timestamp", "")[:10]
                mtype = p.get("media_type", "?")
                lines.append(f"  {i}. ({ts}, {mtype}) {likes} likes, {comments} reacties, bereik {reach} — \"{caption}\"")

        # Posting patronen
        dag_counts = {}
        dagdeel_counts = {}
        for p in month_posts:
            ts = datetime.fromisoformat(p["timestamp"].replace("+0000", "+00:00"))
            dag = DAGEN_NL.get(ts.strftime("%A"), ts.strftime("%A"))
            dag_counts[dag] = dag_counts.get(dag, 0) + 1
            dd = dagdeel(ts.hour)
            dagdeel_counts[dd] = dagdeel_counts.get(dd, 0) + 1
        if dag_counts:
            lines.append(f"\n### Posting patronen:")
            lines.append(f"  Dagen: {dag_counts}")
            lines.append(f"  Dagdelen: {dagdeel_counts}")

    return "\n".join(lines)


def analyze_with_ai(summary: str) -> str:
    """Stuur de samenvatting naar OpenAI GPT-4o-mini voor analyse."""
    from openai import OpenAI

    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return "⚠️  Geen OPENAI_API_KEY gevonden in .env — analyse overgeslagen."

    client = OpenAI(api_key=api_key)

    system_prompt = (
        "Je bent een ervaren social media analist die werkt voor Prins Petfoods, "
        "een Nederlands premium diervoedingsbedrijf. Je analyseert hun Facebook- en "
        "Instagram-prestaties.\n\n"
        "Geef je analyse in het Nederlands met de volgende structuur:\n"
        "1. **Samenvatting** — kort overzicht van de belangrijkste cijfers\n"
        "2. **Sterke punten** — wat gaat goed, welke content scoort\n"
        "3. **Verbeterpunten** — waar liggen kansen\n"
        "4. **Trends** — opvallende patronen in timing, type content, engagement\n"
        "5. **Concrete aanbevelingen** — 3-5 specifieke acties die het team kan ondernemen\n\n"
        "Wees specifiek en actionable. Verwijs naar concrete posts of cijfers. "
        "Vermijd vage algemeenheden zoals 'post meer' of 'wees consistent'. "
        "Houd rekening met de doelgroep: huisdiereigenaren in Nederland/België."
    )

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"Analyseer deze social media data:\n\n{summary}"},
        ],
        temperature=0.7,
        max_tokens=1500,
    )

    return response.choices[0].message.content


def write_analysis(wb, analysis: str):
    """Schrijf de AI-analyse naar een tab 'AI Analyse' in de workbook."""
    sheet_name = "AI Analyse"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(sheet_name)

    now = datetime.now(timezone.utc)
    from openpyxl.styles import Font
    ws["A1"] = "AI Analyse — Prins Petfoods"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Gegenereerd op {now.strftime('%d-%m-%Y %H:%M')} UTC"

    # Schrijf elke regel van de analyse in een eigen rij
    for i, line in enumerate(analysis.split("\n"), start=4):
        ws.cell(row=i, column=1, value=line)

    # Kolombreedte aanpassen
    ws.column_dimensions["A"].width = 120


# ─── Merge CSV + Scraper ─────────────────────────────────────────────

def merge_scraper_data(csv_posts: list[dict], scraper_posts: list[dict]) -> list[dict]:
    """Vul CSV-posts aan met scraper-data waar velden ontbreken."""
    # Index scraper posts op datum (YYYY-MM-DD)
    scraper_by_date = {}
    for p in scraper_posts:
        date_key = p.get("date", "")[:10]
        if date_key:
            scraper_by_date.setdefault(date_key, []).append(p)

    for post in csv_posts:
        date_key = post.get("date", "")[:10]
        matches = scraper_by_date.get(date_key, [])
        if not matches:
            continue
        # Gebruik eerste match op dezelfde dag
        sp = matches[0]
        for field in ("likes", "comments", "shares", "views"):
            if not post.get(field) and sp.get(field):
                post[field] = sp[field]
    return csv_posts


# ─── Main ───────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Social media tracker voor Prins Petfoods")
    parser.add_argument("--no-analysis", action="store_true",
                        help="Sla de AI-analyse over (alleen data ophalen)")
    parser.add_argument("--csv", metavar="MAP",
                        help="Map met CSV exports uit Meta Business Suite")
    args = parser.parse_args()

    wb = load_workbook(EXCEL_FILE)

    prins_fb_posts = []
    edupet_fb_posts = []
    ig_posts = []

    if args.csv:
        # ── CSV Import (primair) ──
        from csv_import import parse_csv_folder
        print(f"\nCSV import uit: {args.csv}")
        csv_data = parse_csv_folder(args.csv)

        for fb_file in csv_data["facebook"]:
            name = fb_file["file"].lower()
            if "edupet" in name:
                edupet_fb_posts = fb_file["posts"]
                print(f"  → Edupet FB: {len(edupet_fb_posts)} posts uit {fb_file['file']}")
            else:
                prins_fb_posts = fb_file["posts"]
                print(f"  → Prins FB: {len(prins_fb_posts)} posts uit {fb_file['file']}")

        for ig_file in csv_data["instagram"]:
            ig_posts = ig_file["posts"]
            print(f"  → Prins IG: {len(ig_posts)} posts uit {ig_file['file']}")

    else:
        # ── Scraper ──
        print("Facebook posts: Prins (scraper)...")
        if has_fb_session():
            result = scrape_fb_page_posts("PrinsPetfoods", max_posts=25, max_scrolls=10)
            prins_fb_posts = result.get("posts", [])
            print(f"  {len(prins_fb_posts)} posts opgehaald via scraper")
        else:
            print("  Geen Facebook-sessie. Draai eerst: python3 fb_scraper.py --login")

        print("Facebook posts: Edupet (scraper)...")
        if has_fb_session():
            result = scrape_fb_page_posts("edupet", max_posts=25, max_scrolls=10)
            edupet_fb_posts = result.get("posts", [])
            print(f"  {len(edupet_fb_posts)} posts opgehaald via scraper")

    # ── Scraper aanvullen als CSV primair is ──
    if args.csv and has_fb_session():
        print("\nScraper: aanvullen ontbrekende data...")
        result = scrape_fb_page_posts("PrinsPetfoods", max_posts=25, max_scrolls=10)
        scraper_posts = result.get("posts", [])
        if scraper_posts and prins_fb_posts:
            prins_fb_posts = merge_scraper_data(prins_fb_posts, scraper_posts)
            print(f"  V {len(scraper_posts)} scraper posts gemerged met Prins FB")

    # ── Schrijf naar Excel ──
    print("\nSchrijven naar Excel...")
    if ig_posts:
        write_ig_posts(wb, ig_posts)
        print("  V Instagram posts geschreven")

    if prins_fb_posts:
        n = write_fb_posts(wb, prins_fb_posts)
        write_fb_kpis(wb, prins_fb_posts)
        print(f"  V Facebook Prins posts geschreven ({n} nieuwe)")
        print("  V Facebook Prins KPI's geschreven")

    if edupet_fb_posts:
        n = write_fb_posts(wb, edupet_fb_posts)
        print(f"  V Facebook Edupet posts geschreven ({n} nieuwe)")

    wb.save(EXCEL_FILE)
    print(f"\nOpgeslagen in {EXCEL_FILE}")

    # ── AI Analyse ──
    if not args.no_analysis:
        print("\n🤖 AI-analyse genereren...\n")
        summary = collect_summary(prins_fb_posts, edupet_fb_posts, ig_posts)
        analysis = analyze_with_ai(summary)
        print("=" * 60)
        print("  SOCIAL MEDIA ANALYSE — Prins Petfoods")
        print("=" * 60)
        print(analysis)
        print("=" * 60)

        # Schrijf analyse naar Excel
        write_analysis(wb, analysis)
        wb.save(EXCEL_FILE)
        print(f"\n✅ Analyse opgeslagen in tab 'AI Analyse' in {EXCEL_FILE}")
    else:
        print("\nAI-analyse overgeslagen (--no-analysis)")


if __name__ == "__main__":
    main()